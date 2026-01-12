"""
Advanced Export Service
Provides Excel, PDF, and Email export functionality
"""

import os
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("ReportLab not available. PDF export will be disabled.")


def export_to_excel(data: List[Dict[str, Any]], filename: str, 
                    sheet_name: str = "Data", format_type: str = "detailed") -> str:
    """
    Export data to Excel with formatting
    
    Args:
        data: List of dictionaries to export
        filename: Output filename
        sheet_name: Name of the Excel sheet
        format_type: Type of formatting (detailed, summary, matches)
    
    Returns:
        Path to exported file
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            ws['A1'] = "No data available"
            wb.save(filename)
            return filename
        
        # Headers
        headers = list(data[0].keys())
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=len(data) + 1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Add border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(data) + 1):
            for cell in row:
                cell.border = thin_border
        
        wb.save(filename)
        logger.info(f"Excel file exported: {filename}")
        return filename
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}", exc_info=True)
        raise


def export_to_pdf(data: List[Dict[str, Any]], filename: str, 
                 title: str = "Report", format_type: str = "detailed") -> Optional[str]:
    """
    Export data to PDF report
    
    Args:
        data: List of dictionaries to export
        filename: Output filename
        title: Report title
        format_type: Type of report (detailed, summary, matches)
    
    Returns:
        Path to exported file or None if ReportLab not available
    """
    if not REPORTLAB_AVAILABLE:
        logger.warning("ReportLab not available. PDF export skipped.")
        return None
    
    try:
        doc = SimpleDocTemplate(filename, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_para = Paragraph(title, styles['Title'])
        elements.append(title_para)
        elements.append(Spacer(1, 0.2 * inch))
        
        # Date
        date_para = Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 0.2 * inch))
        
        if not data:
            no_data = Paragraph("No data available", styles['Normal'])
            elements.append(no_data)
            doc.build(elements)
            return filename
        
        # Table data
        headers = list(data[0].keys())
        table_data = [headers]
        
        for row in data[:100]:  # Limit to 100 rows for PDF
            table_data.append([str(row.get(header, ""))) for header in headers])
        
        if len(data) > 100:
            table_data.append(["...", f"({len(data) - 100} more rows)"])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        logger.info(f"PDF file exported: {filename}")
        return filename
    except Exception as e:
        logger.error(f"Error exporting to PDF: {e}", exc_info=True)
        return None


def export_to_csv(data: List[Dict[str, Any]], filename: str) -> str:
    """
    Export data to CSV
    
    Args:
        data: List of dictionaries to export
        filename: Output filename
    
    Returns:
        Path to exported file
    """
    try:
        df = pd.DataFrame(data)
        df.to_csv(filename, index=False)
        logger.info(f"CSV file exported: {filename}")
        return filename
    except Exception as e:
        logger.error(f"Error exporting to CSV: {e}", exc_info=True)
        raise

